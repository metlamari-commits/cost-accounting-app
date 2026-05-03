#!/usr/bin/env python3
"""
IATE Data Extractor
====================
Εξάγει ορολογικά δεδομένα από το IATE (Inter-Active Terminology for Europe) μέσω:
  1. IATE REST API (https://iate.europa.eu/em-api/entries/_search)
  2. Web scraping (fallback αν δεν υπάρχει API key)
  3. Επεξεργασία TBX/XML αρχείων
Αποθηκεύει σε Excel (.xlsx) και CSV.

Εγκατάσταση dependencies:
    pip install requests beautifulsoup4 lxml openpyxl pandas tqdm

Χρήση:
    # API search
    python iate_extractor.py --mode api --query "artificial intelligence" --lang el,en,fr --output results

    # Web scraping
    python iate_extractor.py --mode scrape --query "machine learning" --lang el,en --output results

    # Επεξεργασία TBX αρχείου
    python iate_extractor.py --mode tbx --tbx-file iate_export.tbx --output results

    # Batch επεξεργασία φακέλου με TBX αρχεία
    python iate_extractor.py --mode tbx --tbx-dir ./tbx_files --output results
"""

import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree as ET

try:
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from tqdm import tqdm
except ImportError as e:
    print(f"[ERROR] Missing dependency: {e}")
    print("Εκτέλεσε: pip install requests beautifulsoup4 lxml openpyxl pandas tqdm")
    sys.exit(1)


# ─────────────────────────────────────────────
# IATE API CLIENT
# ─────────────────────────────────────────────

IATE_API_BASE = "https://iate.europa.eu/em-api"
IATE_SEARCH_URL = f"{IATE_API_BASE}/entries/_search"
IATE_ENTRY_URL = f"{IATE_API_BASE}/entries"

# Γλωσσικοί κωδικοί IATE
LANG_CODES = {
    "bg": "bul", "cs": "ces", "da": "dan", "de": "deu", "el": "ell",
    "en": "eng", "es": "spa", "et": "est", "fi": "fin", "fr": "fra",
    "ga": "gle", "hr": "hrv", "hu": "hun", "it": "ita", "lt": "lit",
    "lv": "lav", "mt": "mlt", "nl": "nld", "pl": "pol", "pt": "por",
    "ro": "ron", "sk": "slk", "sl": "slv", "sv": "swe"
}


def search_iate_api(query: str, languages: list, domain: str = "",
                    api_key: str = "", limit: int = 50, offset: int = 0) -> dict:
    """Αναζήτηση μέσω IATE REST API."""

    # Μετατροπή γλωσσών σε IATE format
    lang_params = []
    for lang in languages:
        code = LANG_CODES.get(lang.lower(), lang.lower())
        lang_params.append(code)

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    payload = {
        "query": query,
        "sourceLanguage": lang_params[0] if lang_params else "eng",
        "targetLanguages": lang_params[1:] if len(lang_params) > 1 else [],
        "domain": domain,
        "offset": offset,
        "limit": limit,
        "includeTerms": True,
        "includeDefinitions": True,
        "includeContexts": True
    }

    try:
        resp = requests.post(IATE_SEARCH_URL, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.HTTPError as e:
        if resp.status_code == 401:
            print("[WARN] Χρειάζεται API key. Χρησιμοποίησε --api-key <key>")
            print("       Εγγραφή: https://iate.europa.eu/about/developer-docs")
        elif resp.status_code == 429:
            print("[WARN] Rate limit. Αναμονή 60 δευτερολέπτων...")
            time.sleep(60)
            return search_iate_api(query, languages, domain, api_key, limit, offset)
        raise
    except Exception as e:
        print(f"[ERROR] API error: {e}")
        return {}


def fetch_all_api_results(query: str, languages: list, domain: str = "",
                           api_key: str = "", max_results: int = 500) -> list:
    """Ανακτά όλα τα αποτελέσματα με pagination."""
    all_entries = []
    offset = 0
    batch_size = 50

    print(f"[API] Αναζήτηση: '{query}' | Γλώσσες: {', '.join(languages)}")

    with tqdm(total=max_results, desc="Ανάκτηση εγγραφών") as pbar:
        while offset < max_results:
            data = search_iate_api(query, languages, domain, api_key,
                                   min(batch_size, max_results - offset), offset)

            if not data or "items" not in data:
                break

            items = data.get("items", [])
            if not items:
                break

            all_entries.extend(items)
            pbar.update(len(items))

            total = data.get("totalItems", 0)
            if offset + len(items) >= total:
                break

            offset += len(items)
            time.sleep(0.5)  # Rate limiting

    print(f"[API] Βρέθηκαν {len(all_entries)} εγγραφές")
    return all_entries


def parse_api_entry(entry: dict, languages: list) -> dict:
    """Μετατρέπει IATE API entry σε flat dict."""
    row = {
        "id": entry.get("id", ""),
        "domain": "",
        "subdomain": "",
        "reliability": entry.get("reliabilityCode", ""),
        "source_lang": "",
        "url": f"https://iate.europa.eu/entry/result/{entry.get('id', '')}"
    }

    # Domain
    domains = entry.get("domain", [])
    if domains:
        row["domain"] = domains[0].get("name", {}).get("en", "")
        if len(domains) > 1:
            row["subdomain"] = domains[1].get("name", {}).get("en", "")

    # Γλωσσικές εγγραφές
    language_entries = entry.get("language", [])
    for lang_entry in language_entries:
        lang_code = lang_entry.get("code", "").lower()

        # Βρες τον σύντομο κωδικό
        short_code = lang_code
        for short, long in LANG_CODES.items():
            if long == lang_code:
                short_code = short
                break

        if languages and short_code not in [l.lower() for l in languages]:
            continue

        terms = lang_entry.get("termEntries", [])
        definitions = lang_entry.get("definitionEntries", [])

        # Κύριος όρος
        main_term = ""
        term_type = ""
        if terms:
            main_term = terms[0].get("value", "")
            term_type = terms[0].get("type", {}).get("name", {}).get("en", "")

        # Ορισμός
        definition = ""
        def_source = ""
        if definitions:
            definition = definitions[0].get("value", "")
            def_source = definitions[0].get("source", "")

        row[f"term_{short_code}"] = main_term
        row[f"type_{short_code}"] = term_type
        row[f"def_{short_code}"] = definition
        row[f"def_source_{short_code}"] = def_source

        # Synonyms
        synonyms = [t.get("value", "") for t in terms[1:] if t.get("value")]
        row[f"synonyms_{short_code}"] = " | ".join(synonyms)

    return row


# ─────────────────────────────────────────────
# WEB SCRAPER
# ─────────────────────────────────────────────

IATE_SEARCH_PAGE = "https://iate.europa.eu/search/result"


def scrape_iate(query: str, languages: list, max_pages: int = 5) -> list:
    """Web scraping από το IATE website."""
    print(f"[SCRAPE] Scraping IATE για: '{query}'")

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (compatible; IATE-Research-Tool/1.0)",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-US,en;q=0.9,el;q=0.8"
    })

    all_entries = []
    src_lang = languages[0].lower() if languages else "en"
    tgt_langs = ",".join(l.lower() for l in languages[1:]) if len(languages) > 1 else "el"

    for page in range(1, max_pages + 1):
        url = (f"{IATE_SEARCH_PAGE}?query={requests.utils.quote(query)}"
               f"&sourceLanguage={src_lang}&targetLanguages={tgt_langs}"
               f"&domain=0&offset={(page-1)*25}&limit=25")

        try:
            resp = session.get(url, timeout=30)
            resp.raise_for_status()
        except Exception as e:
            print(f"[WARN] Σελίδα {page}: {e}")
            break

        soup = BeautifulSoup(resp.content, "lxml")
        entries = _parse_iate_page(soup, languages)

        if not entries:
            print(f"[SCRAPE] Δεν βρέθηκαν αποτελέσματα στη σελίδα {page}")
            break

        all_entries.extend(entries)
        print(f"[SCRAPE] Σελίδα {page}: {len(entries)} εγγραφές")
        time.sleep(1.5)

    print(f"[SCRAPE] Σύνολο: {len(all_entries)} εγγραφές")
    return all_entries


def _parse_iate_page(soup: BeautifulSoup, languages: list) -> list:
    """Parses HTML από IATE αποτελέσματα."""
    entries = []

    # Δοκιμάζουμε διαφορετικά HTML patterns
    result_blocks = (soup.find_all("div", class_="term-entry") or
                     soup.find_all("article", class_="term-result") or
                     soup.find_all("div", attrs={"data-entry-id": True}))

    for block in result_blocks:
        entry = {}

        # ID
        entry["id"] = block.get("data-entry-id", block.get("id", ""))

        # Domain
        domain_el = block.find(class_=["domain", "term-domain"])
        entry["domain"] = domain_el.get_text(strip=True) if domain_el else ""

        # Γλωσσικά blocks
        for lang_block in block.find_all(attrs={"lang": True}):
            lang = lang_block.get("lang", "").lower()[:2]
            if languages and lang not in [l.lower() for l in languages]:
                continue

            term_el = lang_block.find(class_=["term", "headword"])
            entry[f"term_{lang}"] = term_el.get_text(strip=True) if term_el else ""

            def_el = lang_block.find(class_=["definition", "def"])
            entry[f"def_{lang}"] = def_el.get_text(strip=True) if def_el else ""

        if any(k.startswith("term_") for k in entry):
            entries.append(entry)

    # Fallback: αν το HTML structure άλλαξε, χρησιμοποίησε JSON-LD
    if not entries:
        for script in soup.find_all("script", type="application/json"):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list):
                    entries.extend(data)
                elif isinstance(data, dict) and "items" in data:
                    entries.extend(data["items"])
            except (json.JSONDecodeError, AttributeError):
                continue

    return entries


# ─────────────────────────────────────────────
# TBX / XML PARSER
# ─────────────────────────────────────────────

# TBX Namespaces
TBX_NS = {
    "tbx": "urn:iso:std:iso:30042:ed-2",
    "": ""  # default namespace
}


def parse_tbx_file(filepath) -> list:
    """Επεξεργασία TBX (TermBase eXchange) αρχείου."""
    filepath = Path(filepath)
    print(f"[TBX] Επεξεργασία: {filepath.name}")

    entries = []

    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"[ERROR] XML parse error: {e}")
        return []

    # Αναγνώριση TBX namespace
    ns = ""
    tag = root.tag
    if "{" in tag:
        ns = tag[1:tag.index("}")]

    def find(elem, path, ns=""):
        if ns:
            parts = path.split("/")
            ns_path = "/".join(f"{{{ns}}}{p}" if not p.startswith("{") else p
                               for p in parts)
            return elem.find(ns_path)
        return elem.find(path)

    def findall(elem, path, ns=""):
        if ns:
            parts = path.split("/")
            ns_path = "/".join(f"{{{ns}}}{p}" if not p.startswith("{") else p
                               for p in parts)
            return elem.findall(ns_path)
        return elem.findall(path)

    def findtext(elem, path, ns="", default=""):
        result = find(elem, path, ns)
        return result.text.strip() if result is not None and result.text else default

    # Βρες όλα τα conceptEntry / termEntry
    concept_entries = (findall(root, ".//conceptEntry", ns) or
                       findall(root, ".//termEntry", ns) or
                       root.iter(f"{{{ns}}}conceptEntry" if ns else "conceptEntry"))

    for concept in concept_entries:
        row = {}

        # Entry ID
        row["id"] = concept.get("id", "")

        # Subject field (domain)
        for descrip in concept.iter(f"{{{ns}}}descrip" if ns else "descrip"):
            dtype = descrip.get("type", "")
            if "subjectField" in dtype or "domain" in dtype.lower():
                row["domain"] = descrip.text.strip() if descrip.text else ""
                break

        # Γλωσσικές εγγραφές
        lang_sections = (findall(concept, "langSec", ns) or
                         findall(concept, "langSet", ns))

        for lang_sec in lang_sections:
            lang = lang_sec.get("{http://www.w3.org/XML/1998/namespace}lang",
                                lang_sec.get("xml:lang", "")).lower()[:2]

            # Terms
            terms = []
            term_secs = (findall(lang_sec, "termSec", ns) or
                         findall(lang_sec, "tig", ns) or
                         findall(lang_sec, "ntig", ns))

            for term_sec in term_secs:
                term_el = (find(term_sec, "term", ns) or
                           find(term_sec, "termGrp/term", ns))
                if term_el is not None and term_el.text:
                    term_val = term_el.text.strip()

                    # Term type/status
                    term_type = ""
                    for td in term_sec.iter(f"{{{ns}}}termNote" if ns else "termNote"):
                        if "termType" in td.get("type", "") or "normativeAuthorization" in td.get("type", ""):
                            term_type = td.text.strip() if td.text else ""
                            break

                    terms.append({"term": term_val, "type": term_type})

            if terms:
                row[f"term_{lang}"] = terms[0]["term"]
                row[f"type_{lang}"] = terms[0]["type"]
                row[f"synonyms_{lang}"] = " | ".join(
                    t["term"] for t in terms[1:] if t["term"]
                )

            # Definitions
            for descrip in lang_sec.iter(f"{{{ns}}}descrip" if ns else "descrip"):
                if "definition" in descrip.get("type", "").lower():
                    row[f"def_{lang}"] = descrip.text.strip() if descrip.text else ""
                    break

            # Context
            for context in lang_sec.iter(f"{{{ns}}}descrip" if ns else "descrip"):
                if "context" in context.get("type", "").lower():
                    row[f"context_{lang}"] = context.text.strip() if context.text else ""
                    break

        if row:
            entries.append(row)

    print(f"[TBX] Εξήχθηκαν {len(entries)} εγγραφές από {filepath.name}")
    return entries


def parse_tbx_directory(directory) -> list:
    """Batch επεξεργασία φακέλου με TBX αρχεία."""
    directory = Path(directory)
    tbx_files = list(directory.glob("*.tbx")) + list(directory.glob("*.xml"))

    if not tbx_files:
        print(f"[WARN] Δεν βρέθηκαν TBX/XML αρχεία στο: {directory}")
        return []

    print(f"[TBX] Βρέθηκαν {len(tbx_files)} αρχεία")
    all_entries = []

    for tbx_file in tqdm(tbx_files, desc="Επεξεργασία αρχείων"):
        entries = parse_tbx_file(tbx_file)
        for e in entries:
            e["source_file"] = tbx_file.name
        all_entries.extend(entries)

    print(f"[TBX] Σύνολο: {len(all_entries)} εγγραφές")
    return all_entries


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

def entries_to_dataframe(entries: list, languages: list = None) -> pd.DataFrame:
    """Μετατρέπει λίστα entries σε pandas DataFrame."""
    if not entries:
        return pd.DataFrame()

    df = pd.DataFrame(entries)

    # Ταξινόμηση στηλών
    priority_cols = ["id", "domain", "subdomain", "reliability", "url", "source_file"]
    lang_cols = []

    if languages:
        for lang in languages:
            lang = lang.lower()
            for prefix in ["term_", "type_", "synonyms_", "def_", "def_source_", "context_"]:
                col = f"{prefix}{lang}"
                if col in df.columns:
                    lang_cols.append(col)
    else:
        # Auto-detect γλωσσικές στήλες
        for col in df.columns:
            if any(col.startswith(p) for p in ["term_", "type_", "synonyms_", "def_", "context_"]):
                lang_cols.append(col)

    ordered_cols = (
        [c for c in priority_cols if c in df.columns] +
        [c for c in lang_cols if c in df.columns] +
        [c for c in df.columns if c not in priority_cols and c not in lang_cols]
    )

    return df[ordered_cols].fillna("")


def export_to_csv(df: pd.DataFrame, output_path: str):
    """Εξαγωγή σε CSV."""
    path = Path(output_path).with_suffix(".csv")
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"[CSV] Αποθηκεύτηκε: {path} ({len(df)} εγγραφές)")
    return path


def export_to_excel(df: pd.DataFrame, output_path: str, sheet_name: str = "IATE Data"):
    """Εξαγωγή σε Excel με formatting."""
    path = Path(output_path).with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Στυλ
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="003399")  # EU blue
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="EEF3FF")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    # Χρώματα ανά γλώσσα (για visual grouping)
    lang_colors = {
        "el": "FFF9E6", "en": "E6F3FF", "fr": "E6FFE6",
        "de": "FFE6E6", "it": "F3E6FF", "es": "FFE6F3"
    }

    # Headers
    headers = list(df.columns)
    ws.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

        # Ειδικό χρώμα για γλωσσικές στήλες
        for lang, color in lang_colors.items():
            if header.endswith(f"_{lang}"):
                cell.fill = PatternFill("solid",
                                        start_color="003399" if header.startswith("term_") else "1A47A3")
                break

    ws.row_dimensions[1].height = 30

    # Δεδομένα
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        row_vals = list(row_data)
        ws.append(row_vals)

        is_alt = row_idx % 2 == 0

        for col_idx, (header, value) in enumerate(zip(headers, row_vals), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

            # Χρωματισμός ανά γλώσσα
            fill_color = None
            for lang, color in lang_colors.items():
                if header.endswith(f"_{lang}"):
                    fill_color = color if not is_alt else \
                        hex(int(color, 16) - 0x080808)[2:].upper().zfill(6)
                    break

            if fill_color:
                cell.fill = PatternFill("solid", start_color=fill_color)
            elif is_alt:
                cell.fill = alt_fill

            # URL ως hyperlink
            if header == "url" and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(name="Arial", size=10, color="0070C0", underline="single")

            # Wrap για definitions
            if header.startswith("def_") or header.startswith("context_"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Πλάτη στηλών
    col_widths = {
        "id": 12, "domain": 20, "subdomain": 20,
        "reliability": 12, "url": 15, "source_file": 20
    }

    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if header in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[header]
        elif header.startswith("term_"):
            ws.column_dimensions[col_letter].width = 25
        elif header.startswith("def_") or header.startswith("context_"):
            ws.column_dimensions[col_letter].width = 40
        elif header.startswith("synonyms_"):
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 18

    # Freeze panes & auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Sheet με metadata
    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["Πεδίο", "Τιμή"])
    meta_ws.append(["Σύνολο εγγραφών", len(df)])
    meta_ws.append(["Στήλες", ", ".join(headers)])
    meta_ws.append(["Πηγή", "IATE - Inter-Active Terminology for Europe"])
    meta_ws.append(["URL", "https://iate.europa.eu"])
    meta_ws.append(["Εξαγωγή", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")])

    for cell in meta_ws["A"]:
        cell.font = Font(bold=True)

    wb.save(path)
    print(f"[XLSX] Αποθηκεύτηκε: {path} ({len(df)} εγγραφές)")
    return path


# ─────────────────────────────────────────────
# MAIN CLI
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="IATE Data Extractor - Εξαγωγή ορολογίας από IATE",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Παραδείγματα:
  # API αναζήτηση (ελληνικά + αγγλικά)
  python iate_extractor.py --mode api --query "τεχνητή νοημοσύνη" --lang el,en

  # Με API key και domain
  python iate_extractor.py --mode api --query "regulation" --lang en,el,fr \\
      --api-key YOUR_KEY --domain IT --max-results 200

  # Web scraping
  python iate_extractor.py --mode scrape --query "climate change" --lang en,el

  # Ένα TBX αρχείο
  python iate_extractor.py --mode tbx --tbx-file export.tbx

  # Φάκελος με TBX αρχεία
  python iate_extractor.py --mode tbx --tbx-dir ./tbx_exports --output iate_all

  # Μόνο CSV (χωρίς Excel)
  python iate_extractor.py --mode api --query "law" --lang en,el --no-excel
        """
    )

    parser.add_argument("--mode", choices=["api", "scrape", "tbx"], default="api",
                        help="Τρόπος εξαγωγής (default: api)")
    parser.add_argument("--query", "-q", help="Όρος αναζήτησης")
    parser.add_argument("--lang", "-l", default="en,el",
                        help="Γλώσσες διαχωρισμένες με κόμμα (default: en,el)")
    parser.add_argument("--domain", "-d", default="",
                        help="Θεματική περιοχή IATE (π.χ. IT, LAW, ENVIRONMENT)")
    parser.add_argument("--api-key", default="",
                        help="IATE API key (από https://iate.europa.eu)")
    parser.add_argument("--max-results", type=int, default=100,
                        help="Μέγιστος αριθμός αποτελεσμάτων (default: 100)")
    parser.add_argument("--tbx-file", help="Αρχείο TBX για επεξεργασία")
    parser.add_argument("--tbx-dir", help="Φάκελος με TBX αρχεία")
    parser.add_argument("--output", "-o", default="iate_results",
                        help="Όνομα αρχείου εξόδου (χωρίς extension, default: iate_results)")
    parser.add_argument("--no-excel", action="store_true",
                        help="Χωρίς Excel, μόνο CSV")
    parser.add_argument("--no-csv", action="store_true",
                        help="Χωρίς CSV, μόνο Excel")

    args = parser.parse_args()
    languages = [l.strip() for l in args.lang.split(",") if l.strip()]

    print("=" * 60)
    print("  IATE Data Extractor")
    print("=" * 60)

    # Εξαγωγή δεδομένων
    raw_entries = []

    if args.mode == "api":
        if not args.query:
            parser.error("--query απαιτείται για --mode api")

        api_entries = fetch_all_api_results(
            args.query, languages, args.domain,
            args.api_key, args.max_results
        )
        raw_entries = [parse_api_entry(e, languages) for e in api_entries]

    elif args.mode == "scrape":
        if not args.query:
            parser.error("--query απαιτείται για --mode scrape")
        raw_entries = scrape_iate(args.query, languages)

    elif args.mode == "tbx":
        if args.tbx_dir:
            raw_entries = parse_tbx_directory(args.tbx_dir)
        elif args.tbx_file:
            raw_entries = parse_tbx_file(args.tbx_file)
        else:
            parser.error("--tbx-file ή --tbx-dir απαιτείται για --mode tbx")

    if not raw_entries:
        print("[WARN] Δεν βρέθηκαν δεδομένα. Ελέγξτε τις παραμέτρους.")
        sys.exit(0)

    # DataFrame
    df = entries_to_dataframe(raw_entries, languages)
    print(f"\n[INFO] DataFrame: {len(df)} γραμμές × {len(df.columns)} στήλες")

    # Εξαγωγή
    output_dir = Path(args.output).parent
    if not output_dir.exists() and str(output_dir) != ".":
        output_dir.mkdir(parents=True)

    if not args.no_csv:
        export_to_csv(df, args.output)

    if not args.no_excel:
        export_to_excel(df, args.output)

    print("\n✓ Ολοκληρώθηκε!")


if __name__ == "__main__":
    main()
