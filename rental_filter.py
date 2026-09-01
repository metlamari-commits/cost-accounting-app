import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Ρυθμίσεις
MAX_RENT = 500
TARGET_AREAS = ["ωρωπός", "κάλαμος", "άγιοι απόστολοι"]

# Προσεγγιστικές αποστάσεις από Μαλακάσα (σε km)
DISTANCES_FROM_MALAKASA = {
    "ωρωπός": 10,
    "κάλαμος": 8,
    "άγιοι απόστολοι": 7,
}

# ΕΔΩ βάζεις τις αγγελίες που βρίσκεις (χειροκίνητα από site)
# title: περιγραφή, price: τιμή σε €, area: περιοχή
listings = [
    {"title": "Διαμέρισμα 40τμ στον Ωρωπό, 450€ χωρίς μεσίτη", "price": 450, "area": "ωρωπός"},
    {"title": "Γκαρσονιέρα 30τμ στον Κάλαμο, 520€", "price": 520, "area": "κάλαμος"},
    {"title": "Στούντιο στους Αγίους Αποστόλους, 400€ χωρίς μεσίτη", "price": 400, "area": "άγιοι απόστολοι"},
    # πρόσθεσε κι άλλες εδώ...
]


def filter_listings(listings):
    results = []
    for item in listings:
        area = item["area"].lower()
        price = item["price"]

        if area not in TARGET_AREAS:
            continue

        if price > MAX_RENT:
            continue

        distance = DISTANCES_FROM_MALAKASA.get(area)
        item["distance_km"] = distance
        results.append(item)

    return sorted(results, key=lambda x: x.get("distance_km", 999))


def export_to_excel(filtered, filename="αποτελέσματα_ενοικίων.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Αγγελίες"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["Τίτλος", "Τιμή (€)", "Περιοχή", "Απόσταση από Μαλακάσα (km)"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    for row_idx, item in enumerate(filtered, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        values = [
            item["title"],
            item["price"],
            item["area"].title(),
            item.get("distance_km", "N/A"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center")

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.row_dimensions[1].height = 20

    wb.save(filename)
    print(f"Excel αποθηκεύτηκε: {filename}")


def main():
    filtered = filter_listings(listings)

    print("\nΑΠΟΤΕΛΕΣΜΑΤΑ ΕΩΣ 500€ ΣΕ ΩΡΩΠΟ - ΚΑΛΑΜΟ - ΑΓΙΟΥΣ ΑΠΟΣΤΟΛΟΥΣ\n")
    for item in filtered:
        print(f"- {item['title']}")
        print(f"  Τιμή: {item['price']}€")
        print(f"  Περιοχή: {item['area']}")
        print(f"  Απόσταση από Μαλακάσα: {item.get('distance_km', 'N/A')} km\n")

    export_to_excel(filtered)


if __name__ == "__main__":
    main()
